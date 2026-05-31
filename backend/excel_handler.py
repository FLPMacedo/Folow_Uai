"""Import/export de clientes em .xlsx.

Formato planilha (linha 1 = cabeçalho, demais = dados):

    nome | telefone | email | data_nascimento | data_inicio_parceria |
    plano | grupo | status | observacoes | tags

Datas aceitas: nativo Excel, "DD/MM/AAAA", "AAAA-MM-DD".
Tags: lista separada por vírgula. Status: 'ativo'/'inativo' (default 'ativo').

Uso:
    from backend.excel_handler import import_clientes_xlsx, export_clientes_xlsx
    rows, errors = import_clientes_xlsx("planilha.xlsx")
    export_clientes_xlsx(clientes, "out.xlsx")
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import Workbook, load_workbook


COLUMNS: list[str] = [
    "nome",
    "telefone",
    "email",
    "data_nascimento",
    "data_inicio_parceria",
    "plano",
    "grupo",
    "status",
    "observacoes",
    "tags",
]
REQUIRED: set[str] = {"nome", "telefone", "data_inicio_parceria"}
VALID_STATUS: set[str] = {"ativo", "inativo"}


# ============================================================================
# Erros
# ============================================================================
@dataclass(frozen=True)
class ImportError:  # noqa: A001 — shadowing builtin é intencional dentro do módulo
    row: int  # 1-based, mesma numeração do Excel (cabeçalho = 1)
    column: Optional[str]
    message: str

    def __str__(self) -> str:
        col = f" [{self.column}]" if self.column else ""
        return f"linha {self.row}{col}: {self.message}"


# ============================================================================
# Parse helpers
# ============================================================================
def _parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        raise ValueError(f"data inválida: {value!r}")
    text = value.strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"data inválida: {value!r} (use DD/MM/AAAA ou AAAA-MM-DD)")


def _clean_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _has_min_digits(phone: str, n: int = 10) -> bool:
    return sum(1 for c in phone if c.isdigit()) >= n


def _parse_tags(value: Any) -> list[str]:
    s = _clean_str(value)
    if not s:
        return []
    return [t.strip() for t in s.split(",") if t.strip()]


# ============================================================================
# Template generator
# ============================================================================
def generate_template_xlsx(path: str | Path) -> Path:
    """Cria planilha vazia com cabeçalhos + 1 linha exemplo."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(COLUMNS)
    ws.append([
        "Maria Silva",
        "+5531999990001",
        "maria@example.com",
        "15/05/1990",
        "30/05/2025",
        "Premium",
        "VIP",
        "ativo",
        "Cliente desde 2025",
        "VIP,corredora",
    ])
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(
            14, len(col) + 2
        )
    wb.save(p)
    return p


# ============================================================================
# Import
# ============================================================================
def import_clientes_xlsx(
    path: str | Path,
) -> tuple[list[dict[str, Any]], list[ImportError]]:
    """Lê .xlsx, retorna (linhas válidas como dict, erros por linha).

    Linhas com erro são puladas mas NÃO abortam o import — a UI mostra a lista.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {p}")

    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [ImportError(row=1, column=None, message="planilha vazia")]

    header = [_clean_str(c) for c in header_row]
    missing = REQUIRED - {c for c in header if c}
    if missing:
        wb.close()
        return [], [
            ImportError(row=1, column=None,
                        message=f"colunas obrigatórias ausentes: {sorted(missing)}")
        ]

    col_index = {name: i for i, name in enumerate(header) if name in COLUMNS}

    valid: list[dict[str, Any]] = []
    errors: list[ImportError] = []

    for idx, raw in enumerate(rows_iter, start=2):
        if raw is None or all(c is None or (isinstance(c, str) and not c.strip())
                              for c in raw):
            continue  # linha em branco

        row_errors: list[ImportError] = []
        record: dict[str, Any] = {}

        def take(name: str) -> Any:
            i = col_index.get(name)
            if i is None or i >= len(raw):
                return None
            return raw[i]

        # nome
        nome = _clean_str(take("nome"))
        if not nome:
            row_errors.append(ImportError(idx, "nome", "obrigatório"))
        record["nome"] = nome

        # telefone
        tel = _clean_str(take("telefone"))
        if not tel:
            row_errors.append(ImportError(idx, "telefone", "obrigatório"))
        elif not _has_min_digits(tel, 10):
            row_errors.append(ImportError(idx, "telefone",
                                          f"poucos dígitos: {tel!r}"))
        record["telefone"] = tel

        # email
        record["email"] = _clean_str(take("email"))

        # data_nascimento (opc)
        try:
            record["data_nascimento"] = _parse_date(take("data_nascimento"))
        except ValueError as e:
            row_errors.append(ImportError(idx, "data_nascimento", str(e)))

        # data_inicio_parceria (obrig)
        try:
            d = _parse_date(take("data_inicio_parceria"))
            if d is None:
                row_errors.append(ImportError(idx, "data_inicio_parceria",
                                              "obrigatório"))
            record["data_inicio_parceria"] = d
        except ValueError as e:
            row_errors.append(ImportError(idx, "data_inicio_parceria", str(e)))

        record["plano"] = _clean_str(take("plano"))
        record["grupo"] = _clean_str(take("grupo"))

        status = _clean_str(take("status")) or "ativo"
        if status.lower() not in VALID_STATUS:
            row_errors.append(ImportError(idx, "status",
                                          f"valor inválido {status!r}; use 'ativo' ou 'inativo'"))
        record["status"] = status.lower()

        record["observacoes"] = _clean_str(take("observacoes"))
        record["tags"] = _parse_tags(take("tags"))

        if row_errors:
            errors.extend(row_errors)
        else:
            valid.append(record)

    wb.close()
    return valid, errors


# ============================================================================
# Export
# ============================================================================
def export_clientes_xlsx(
    clientes: Iterable[Any],
    path: str | Path,
) -> Path:
    """Escreve clientes em .xlsx no mesmo schema do import.

    Aceita: dict, SQLModel Cliente, ou qualquer objeto com atributos das COLUMNS.
    Campo `tags` aceita list[str] ou list[Tag] (usa .nome).
    """
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(COLUMNS)

    for c in clientes:
        ws.append([_pull(c, col) for col in COLUMNS])

    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(
            14, len(col) + 2
        )
    wb.save(p)
    return p


def _pull(obj: Any, key: str) -> Any:
    """Extrai campo do objeto/dict. Tags vira CSV. Datas string ISO."""
    if isinstance(obj, dict):
        val = obj.get(key)
    else:
        val = getattr(obj, key, None)
    if key == "tags":
        if not val:
            return ""
        return ",".join(
            (t if isinstance(t, str) else getattr(t, "nome", str(t))) for t in val
        )
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if val is None:
        return ""
    if hasattr(val, "value"):  # enum
        return val.value
    return val
