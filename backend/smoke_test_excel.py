"""Smoke test offline para excel_handler.

Run:
    python -m backend.smoke_test_excel

Verifica:
  1. generate_template_xlsx cria arquivo, cabeçalho correto
  2. import_clientes_xlsx lê template (1 linha exemplo)
  3. Linhas válidas e inválidas — erros são acumulados, válidas passam
  4. Date variants: DD/MM/AAAA, AAAA-MM-DD, datetime nativo Excel
  5. Tags CSV → lista limpa
  6. Required columns check
  7. Roundtrip: import → export → import devolve mesmos dados
  8. Export aceita dict, objetos com atributos e Cliente SQLModel
"""
from __future__ import annotations

import sys
import tempfile
import traceback
from datetime import date, datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook  # noqa: E402

from backend.excel_handler import (  # noqa: E402
    COLUMNS,
    export_clientes_xlsx,
    generate_template_xlsx,
    import_clientes_xlsx,
)
from backend.models import Cliente, StatusCliente  # noqa: E402


results: list[tuple[str, bool, str]] = []


def check(label: str, ok: bool, detail: str = "") -> None:
    results.append((label, ok, detail))
    icon = "OK " if ok else "FAIL"
    print(f"[{icon}] {label}" + (f" — {detail}" if detail else ""))


def _write_test_sheet(path: Path, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(COLUMNS)
    for r in rows:
        ws.append(r)
    wb.save(path)


def main() -> int:
    td = Path(tempfile.mkdtemp(prefix="followuai-smoke-"))
    print(f"Tempdir: {td}")

    # =====================================================================
    # 1. generate_template
    # =====================================================================
    tpl = generate_template_xlsx(td / "template.xlsx")
    check("Template arquivo criado", tpl.exists())
    wb = load_workbook(tpl, data_only=True, read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    wb.close()
    check("Cabeçalho bate com COLUMNS", header == COLUMNS,
          f"got={header}")

    # =====================================================================
    # 2. ler template (tem 1 linha exemplo)
    # =====================================================================
    rows, errs = import_clientes_xlsx(tpl)
    check("Template tem 1 linha exemplo válida",
          len(rows) == 1 and len(errs) == 0,
          f"rows={len(rows)} errs={len(errs)}")
    if rows:
        r = rows[0]
        check("Exemplo: nome",            r["nome"] == "Maria Silva", f"{r['nome']!r}")
        check("Exemplo: telefone",        r["telefone"] == "+5531999990001", f"{r['telefone']!r}")
        check("Exemplo: data DD/MM/AAAA → date",
              r["data_nascimento"] == date(1990, 5, 15),
              f"{r['data_nascimento']!r}")
        check("Exemplo: status default 'ativo'", r["status"] == "ativo")
        check("Exemplo: tags CSV → list",
              r["tags"] == ["VIP", "corredora"], f"{r['tags']!r}")

    # =====================================================================
    # 3. linhas válidas e inválidas misturadas
    # =====================================================================
    sample = td / "sample.xlsx"
    _write_test_sheet(sample, [
        # linha 2: válida
        ["João",      "+5531977770001", "j@x.com", "01/01/1985",
         "10/03/2024", "Plano A", "Grupo1", "ativo", "", ""],
        # linha 3: válida (datetime nativo)
        ["Ana",       "+5531966660002", None,       datetime(1992, 7, 20),
         datetime(2023, 1, 1), None, None, "inativo", "obs", "vip"],
        # linha 4: data inválida (deve falhar)
        ["Carlos",    "+5531955550003", "c@x.com",  "32/13/1990",
         "01/01/2024", None, None, "ativo", "", ""],
        # linha 5: telefone faltando (deve falhar)
        ["Pedro",     "",               None,       None,
         "01/01/2024", None, None, "ativo", "", ""],
        # linha 6: nome faltando + status inválido (deve falhar 2x)
        ["",          "+5531944440005", None,       None,
         "01/01/2024", None, None, "MAYBE", "", ""],
        # linha 7: tudo branco — pulada
        [None, None, None, None, None, None, None, None, None, None],
        # linha 8: data_inicio faltando (obrig)
        ["Luiza", "+5531933330007", None, None, None, None, None, "ativo", "", ""],
    ])
    rows, errs = import_clientes_xlsx(sample)
    check("Sample: 2 linhas válidas (João, Ana)",
          len(rows) == 2 and {r["nome"] for r in rows} == {"João", "Ana"},
          f"rows={[r['nome'] for r in rows]}")

    err_rows = {(e.row, e.column) for e in errs}
    expected_err_keys = {
        (4, "data_nascimento"),       # 32/13/1990
        (5, "telefone"),              # vazio
        (6, "nome"),                  # vazio
        (6, "status"),                # MAYBE
        (8, "data_inicio_parceria"),  # vazio (obrig)
    }
    check("Sample: erros nas linhas certas",
          expected_err_keys.issubset(err_rows),
          f"missing={expected_err_keys - err_rows} got={sorted(err_rows)}")
    # garantir que nenhuma linha em branco gerou erro
    check("Linha em branco pulada (sem erro)",
          not any(e.row == 7 for e in errs),
          f"errs em 7={[str(e) for e in errs if e.row == 7]}")

    # =====================================================================
    # 4. date variants — ISO e DD-MM
    # =====================================================================
    dates = td / "dates.xlsx"
    _write_test_sheet(dates, [
        ["A", "+5531900000001", None, "2000-06-15", "2024-01-01",
         None, None, None, None, None],
        ["B", "+5531900000002", None, "15-06-2000", "01-01-2024",
         None, None, None, None, None],
    ])
    rows, errs = import_clientes_xlsx(dates)
    check("Date ISO YYYY-MM-DD parseado",
          len(rows) >= 1 and rows[0]["data_nascimento"] == date(2000, 6, 15),
          f"rows[0]={rows[0] if rows else None}")
    check("Date DD-MM-YYYY parseado",
          len(rows) == 2 and rows[1]["data_inicio_parceria"] == date(2024, 1, 1),
          f"errs={[str(e) for e in errs]}")

    # =====================================================================
    # 5. required column missing
    # =====================================================================
    bad_hdr = td / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["nome", "email"])  # faltando telefone e data_inicio
    ws.append(["A", "a@x.com"])
    wb.save(bad_hdr)
    rows, errs = import_clientes_xlsx(bad_hdr)
    check("Colunas obrig ausentes → erro fatal",
          len(rows) == 0 and len(errs) == 1
          and "obrigatórias" in errs[0].message,
          f"errs={[str(e) for e in errs]}")

    # =====================================================================
    # 6. roundtrip: import → export → import
    # =====================================================================
    rows_in, _ = import_clientes_xlsx(sample)
    out = td / "roundtrip.xlsx"
    export_clientes_xlsx(rows_in, out)
    check("Export arquivo criado", out.exists())
    rows_back, errs_back = import_clientes_xlsx(out)
    check("Roundtrip sem erros", not errs_back,
          f"errs={[str(e) for e in errs_back]}")
    check("Roundtrip: mesma contagem", len(rows_back) == len(rows_in),
          f"in={len(rows_in)} back={len(rows_back)}")
    if rows_back:
        nomes_in = {r["nome"] for r in rows_in}
        nomes_back = {r["nome"] for r in rows_back}
        check("Roundtrip: mesmos nomes", nomes_in == nomes_back,
              f"in={nomes_in} back={nomes_back}")
        # tags devem voltar como lista
        ana_in = next(r for r in rows_in if r["nome"] == "Ana")
        ana_back = next(r for r in rows_back if r["nome"] == "Ana")
        check("Roundtrip: tags preservadas",
              ana_in["tags"] == ana_back["tags"],
              f"in={ana_in['tags']} back={ana_back['tags']}")

    # =====================================================================
    # 7. export aceita SQLModel Cliente
    # =====================================================================
    c = Cliente(
        nome="ORM Test",
        telefone="+5531988880099",
        data_inicio_parceria=date(2024, 6, 1),
        status=StatusCliente.ativo,
    )
    out2 = td / "from_orm.xlsx"
    export_clientes_xlsx([c], out2)
    rows_orm, _ = import_clientes_xlsx(out2)
    check("Export SQLModel Cliente → import OK",
          len(rows_orm) == 1 and rows_orm[0]["nome"] == "ORM Test"
          and rows_orm[0]["status"] == "ativo",
          f"got={rows_orm}")

    # =====================================================================
    # summary
    # =====================================================================
    passed = sum(1 for _, ok, _ in results if ok)
    total = len(results)
    print(f"\n{'='*60}\n{passed}/{total} checks passed\n{'='*60}")
    if passed != total:
        print("\nFailures:")
        for label, ok, detail in results:
            if not ok:
                print(f"  - {label}: {detail}")
        return 1

    # cleanup
    import shutil
    shutil.rmtree(td, ignore_errors=True)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception:
        traceback.print_exc()
        raise SystemExit(2)
